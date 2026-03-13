import streamlit as st
import pandas as pd
import qrcode
from datetime import datetime
import os
import uuid
from io import BytesIO

# Cấu hình file
FILE_NAME = "danh_sach_khach.xlsx"

# 1. KHỞI TẠO FILE NẾU CHƯA CÓ
if not os.path.exists(FILE_NAME):
    df = pd.DataFrame(columns=["ID", "HoTen", "SDT", "GapAi", "MucDich", "GioVao", "GioRa"])
    df.to_excel(FILE_NAME, index=False)

st.set_page_config(page_title="Hệ thống QR Khách", layout="centered")

# --- PHẦN XỬ LÝ CHECK-OUT QUA QR  ---
# --- ĐOẠN XỬ LÝ KHI KHÁCH QUÉT QR RA VỀ ---
params = st.query_params
if "action" in params and params["action"] == "checkout":
    target_id = params.get("id")
    
    # Hiển thị tiêu đề trang trọng
    st.title("🚀 XÁC NHẬN RA VỀ")
    
    if os.path.exists(FILE_NAME):
        df = pd.read_excel(FILE_NAME)
        # Kiểm tra ID và trạng thái
        mask = (df['ID'].astype(str) == target_id) & (df['GioRa'].isna() | (df['GioRa'] == ""))
        
        if mask.any():
            # Lấy tên khách để chào cho thân thiện
            ten_khach = df.loc[mask, 'HoTen'].values[0]
            
            # Ghi giờ ra
            df.loc[mask, 'GioRa'] = datetime.now().strftime("%H:%M %d/%m/%Y")
            df.to_excel(FILE_NAME, index=False)
            
            # --- HIỂN THỊ THÔNG BÁO CẢM ƠN SIÊU ĐẸP ---
            st.success(f"✅ Xác nhận thành công cho khách: **{ten_khach}**")
            
            # Tạo một khung thông báo lớn và trang trọng
            st.markdown(f"""
                <div style="background-color: #f0f2f6; padding: 30px; border-radius: 15px; border-left: 10px solid #00a010; margin-top: 20px;">
                    <h2 style="color: #008000; text-align: center;">🙏 CẢM ƠN QUÝ KHÁCH ĐÃ GHÉ THĂM</h2>
                    <h3 style="color: #444; text-align: center;">HẸN GẶP LẠI!</h3>
                    <p style="text-align: center; font-style: italic;">Hệ thống đã ghi nhận giờ ra về của bạn vào lúc: {datetime.now().strftime("%H:%M")}</p>
                </div>
            """, unsafe_allow_index=True)
            
            # Bắn pháo hoa chúc mừng khách ra về vui vẻ
            st.balloons() 
            
        else:
            # Kiểm tra xem có phải đã báo ra rồi không
            already_out = (df['ID'].astype(str) == target_id) & (df['GioRa'].notna())
            if already_out.any():
                st.info("💡 Bạn đã thực hiện xác nhận ra về trước đó rồi. Chúc bạn một ngày tốt lành!")
            else:
                st.error("❌ Mã định danh không hợp lệ hoặc không tồn tại trong hệ thống.")
    
    # Nút quay lại trang chủ (nếu cần)
    if st.button("Quay lại trang đăng ký"):
        st.query_params.clear()
        st.rerun()
        
    st.stop() # Dừng tại đây để khách không thấy phần của bảo vệ

# --- GIAO DIỆN CHÍNH ---
st.sidebar.title("🔑 QUẢN TRỊ")
user_role = st.sidebar.selectbox("Bạn là ai?", ["Khách hàng", "Bảo vệ / Admin"])

if user_role == "Khách hàng":
    st.title("📝 ĐĂNG KÝ VÀO CƠ QUAN")
    name = st.text_input("Họ và tên")
    phone = st.text_input("Số điện thoại")
    bo_phan = st.selectbox("Bộ phận cần gặp", ["Ban giám hiệu", "Hành chính", "Kế toán", "Khác"])
    
    muc_dich = ""
    if bo_phan == "Hành chính":
        muc_dich = st.text_area("🎯 Mục đích công việc (Bắt buộc):")
    
    if st.button("Lấy mã QR"):
        if name and phone:
            if bo_phan == "Hành chính" and not muc_dich:
                st.error("Vui lòng nhập mục đích!")
            else:
                new_id = str(uuid.uuid4())[:8]
                df_curr = pd.read_excel(FILE_NAME)
                new_row = {
                    "ID": new_id, "HoTen": name, "SDT": phone, 
                    "GapAi": bo_phan, "MucDich": muc_dich, 
                    "GioVao": datetime.now().strftime("%H:%M %d/%m"), "GioRa": ""
                }
                df_curr = pd.concat([df_curr, pd.DataFrame([new_row])], ignore_index=True)
                df_curr.to_excel(FILE_NAME, index=False)
                
                # Tạo QR (Thay link của em vào đây nhé)
                link_goc = "https://he-thong-quan-ly-khach-ra-vao.streamlit.app/" 
                qr_img = qrcode.make(f"{link_goc}?action=checkout&id={new_id}")
                buf = BytesIO()
                qr_img.save(buf, format="PNG")
                st.image(buf.getvalue(), caption="Quét mã này khi ra về", width=300)
        else:
            st.error("Vui lòng nhập đủ tên và SĐT!")

# --- PHẦN DÀNH CHO BẢO VỆ (Giữ nguyên logic cũ, chỉ nâng cấp xuất file) ---
# ... (Phần trên là code đăng ký của khách, em giữ nguyên) ...

# --- PHẦN DÀNH CHO QUẢN TRỊ (BẢO VỆ) - BẢN TOÀN NĂNG ---
else:
    st.title("🛡️ KHU VỰC QUẢN TRỊ")
    password = st.text_input("Nhập mật khẩu quản lý", type="password")
    
    if password == "123456":
        st.success("🛡️ Đã đăng nhập quyền Bảo vệ/Quản trị")
        
        # Đọc dữ liệu từ file Excel
        if os.path.exists(FILE_NAME):
            df = pd.read_excel(FILE_NAME)
            
            # --- BẢNG 1: KHÁCH ĐANG Ở TRONG ---
            st.subheader("🔴 Khách đang ở trong cơ quan")
            # Lọc khách có Giờ Ra trống
            khach_trong = df[df['GioRa'].isna() | (df['GioRa'] == "")]
            if not khach_trong.empty:
                st.dataframe(khach_trong, use_container_width=True)
            else:
                st.info("Hiện không có khách nào ở trong.")

            st.divider()

            # --- BẢNG 2: KHÁCH ĐÃ RA VỀ ---
            st.subheader("🟢 Khách đã ra về trong ngày")
            # Lọc khách đã có Giờ Ra
            khach_ve = df[df['GioRa'].notna() & (df['GioRa'] != "")]
            if not khach_ve.empty:
                # Hiện danh sách đảo ngược (người về mới nhất lên đầu)
                st.dataframe(khach_ve.iloc[::-1], use_container_width=True)
            else:
                st.info("Chưa có khách nào báo ra về.")

            st.divider()

            # --- PHẦN XUẤT FILE EXCEL "PHONG CÁCH ĐIỆN THOẠI" ---
            st.subheader("📝 Công cụ báo cáo")
            
            from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl import Workbook
            import io

            # Nút bấm để chuẩn bị file (Giúp Mobile chạy ổn định)
            if st.button("📊 Chuẩn bị file Excel (Bản đẹp)"):
                ngay_hien_tai = datetime.now().strftime("%d_%m_%Y")
                buffer = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "BaoCaoRaVao"

                headers = list(df.columns)
                ws.append(headers)

                for r in df.values.tolist():
                    row_data = [str(x) if str(x) != 'nan' else "" for x in r]
                    ws.append(row_data)
                
                # --- ĐỊNH DẠNG XANH + KẺ BẢNG ---
                blue_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = blue_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center')

                # Tự chỉnh độ rộng cột
                column_widths = [12, 25, 15, 20, 35, 20, 20] 
                for i, width in enumerate(column_widths):
                    if i < len(headers):
                        ws.column_dimensions[get_column_letter(i + 1)].width = width

                wb.save(buffer)
                
                # Nút tải thực sự hiện ra sau khi chuẩn bị xong
                st.download_button(
                    label="📥 Tải file về máy (Click để lưu)",
                    data=buffer.getvalue(),
                    file_name=f"Bao_cao_khach_{ngay_hien_tai}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            # Nút xóa dữ liệu (Cẩn thận khi dùng)
            st.divider()
            if st.button("🗑️ Reset dữ liệu ngày mới"):
                df_reset = pd.DataFrame(columns=["ID", "HoTen", "SDT", "GapAi", "MucDich", "GioVao", "GioRa"])
                df_reset.to_excel(FILE_NAME, index=False)
                st.warning("Đã làm sạch dữ liệu. Vui lòng F5 (Reload) trang web.")
        else:
            st.error("Chưa có dữ liệu khách đăng ký!")
            
    elif password != "":
        st.error("Mật khẩu không chính xác!")





